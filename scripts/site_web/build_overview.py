#!/usr/bin/env python3
"""Regenere la feuille Excel Vue d'ensemble depuis les feuilles Jour N."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from outils.excel_utils import (
    backup_excel,
    cell_value,
    data_dir,
    day_sheets,
    default_excel_path,
    iter_activity_rows,
    jour_color,
    normalize_text,
    parse_float,
    parse_prix,
)
from outils.overview_config import default_overview_config_path, resolve_overview_config

MOIS_FR = (
    "",
    "janvier",
    "février",
    "mars",
    "avril",
    "mai",
    "juin",
    "juillet",
    "août",
    "septembre",
    "octobre",
    "novembre",
    "décembre",
)

BANNER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
BANNER_FONT = Font(bold=True, size=15, color="FFFFFF")
SUBTITLE_FONT = Font(size=10, color="5D6D7E", italic=True)
AUTO_GENERATED_FONT = Font(size=9, color="95A5A6", italic=True)
TABLE_HEADER_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
TABLE_HEADER_FONT = Font(bold=True, size=10, color="2C3E50")
SECTION_HEADER_FILL = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
SECTION_HEADER_FONT = Font(bold=True, size=11, color="2C3E50")
HIGHLIGHT_FILL = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
BODY_FONT = Font(size=10)
JOUR_FONT = Font(bold=True, size=11, color="FFFFFF")
LODGING_FONT = Font(size=10, color="1A5276")
LINK_FONT = Font(size=10, color="0563C1", underline="single")
TOTAL_FONT = Font(bold=True, size=10, color="2C3E50")

THIN_SIDE = Side(style="thin", color="BDC3C7")
TABLE_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=Side(style="medium", color="7F8C8D"),
)

OVERVIEW_COL_WIDTHS = (10, 28, 32, 40, 11, 12)
OVERVIEW_MAX_COL = 6
DATA_ROW_HEIGHT = 22
PLACEHOLDER_MARKERS = ("à compléter", "a completer", "(à compléter)")


def is_trajet_line(nom: str) -> bool:
    lower = normalize_text(nom).lower()
    return lower.startswith(("trajet ", "retour "))


def parse_start_date(value: str) -> date:
    match = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", value.strip())
    if not match:
        raise ValueError(f"Date invalide {value!r} (attendu AAAA-MM-JJ)")
    return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))


def format_date_fr(d: date) -> str:
    return f"{d.day} {MOIS_FR[d.month]} {d.year}"


def format_date_range_fr(start: date, end: date) -> str:
    if start == end:
        return format_date_fr(start)
    if start.year == end.year and start.month == end.month:
        return f"{start.day} au {end.day} {MOIS_FR[start.month]} {start.year}"
    if start.year == end.year:
        return f"{start.day} {MOIS_FR[start.month]} au {end.day} {MOIS_FR[end.month]} {start.year}"
    return f"{format_date_fr(start)} au {format_date_fr(end)}"


def date_for_jour(start: date, jour: int) -> date:
    return start + timedelta(days=jour - 1)


def parse_heure(value: Any) -> int | None:
    text = normalize_text(value).lower().replace(" ", "")
    match = re.match(r"^(\d{1,2})h(?:\d{2})?$", text)
    if not match:
        return None
    return int(match.group(1))


def is_overnight_stay(row: dict[str, Any]) -> bool:
    """Une nuit = arrivée le soir ; exclut le départ/check-out le matin."""
    if not is_hebergement(row["action"]):
        return False
    ouverture = parse_heure(row.get("ouverture"))
    if ouverture is not None and ouverture < 12:
        return False
    return True


def format_lodging_dates_label(start: date, end: date, nights: int) -> str:
    label = format_date_range_fr(start, end)
    if nights > 1:
        return f"{label} ({nights} nuits)"
    if nights == 1:
        return f"{label} (1 nuit)"
    return label


def row_entry(item: dict[str, Any]) -> dict[str, Any]:
    row = item["row"]
    ci = item["col_index"]
    nom = item["nom"]
    return {
        "jour": item["jour"],
        "visite": item["visite"],
        "nom": nom,
        "ville": normalize_text(cell_value(row, ci, "Ville")),
        "action": normalize_text(cell_value(row, ci, "Action")),
        "type": normalize_text(cell_value(row, ci, "Type")),
        "prix": parse_prix(cell_value(row, ci, "Prix")),
        "remarque": normalize_text(cell_value(row, ci, "Remarque")),
        "ouverture": normalize_text(cell_value(row, ci, "Ouverture")),
        "fermeture": normalize_text(cell_value(row, ci, "Fermeture")),
        "lat": parse_float(cell_value(row, ci, "Latitude")),
        "lon": parse_float(cell_value(row, ci, "Longitude")),
        "lien": normalize_text(cell_value(row, ci, "Lien")),
        "site": normalize_text(cell_value(row, ci, "Site")),
        "is_trajet_line": is_trajet_line(nom),
    }


def is_hebergement(action: str) -> bool:
    return normalize_text(action).lower() in ("hébergement", "hebergement")


def lodging_ville(row: dict[str, Any]) -> str:
    """Ville réelle de la nuit (domicile Ennevelin même si la colonne Ville indique Lille)."""
    nom = normalize_text(row.get("nom", "")).lower()
    if "ennevelin" in nom:
        return "Ennevelin"
    return normalize_text(row.get("ville", ""))


def is_placeholder(value: str) -> bool:
    lower = normalize_text(value).lower()
    return any(marker in lower for marker in PLACEHOLDER_MARKERS)


def format_budget_cell(prix: float, activities: int) -> str:
    if activities == 0:
        return "—"
    if not prix:
        return "0 €"
    text = f"{prix:g}".replace(".", ",")
    return f"~{text} €"


def excel_color(hex_color: str) -> str:
    return hex_color.lstrip("#").upper()


def lighten_hex(hex_color: str, white_blend: float = 0.88) -> str:
    color = excel_color(hex_color)
    red = int(color[0:2], 16)
    green = int(color[2:4], 16)
    blue = int(color[4:6], 16)
    red = int(red + (255 - red) * white_blend)
    green = int(green + (255 - green) * white_blend)
    blue = int(blue + (255 - blue) * white_blend)
    return f"{red:02X}{green:02X}{blue:02X}"


def fill_from_hex(hex_color: str, *, light: bool = False) -> PatternFill:
    color = lighten_hex(hex_color) if light else excel_color(hex_color)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def nightly_lodging_by_jour(rows: list[dict[str, Any]]) -> dict[int, dict[str, str]]:
    """Dernière arrivée (nuit) par jour — exclut les check-out matinaux."""
    by_jour: dict[int, dict[str, str]] = {}
    for row in sorted(rows, key=lambda r: (r["jour"], r["visite"])):
        if not is_overnight_stay(row):
            continue
        nom = row["nom"] or "À compléter"
        ville = lodging_ville(row)
        by_jour[row["jour"]] = {"nom": nom, "ville": ville}
    return by_jour


def format_night_cell(entry: dict[str, str] | None) -> str:
    if not entry:
        return "—"
    nom = entry.get("nom") or ""
    ville = entry.get("ville") or ""
    if nom and ville and ville.lower() not in nom.lower():
        return f"{nom} ({ville})"
    return nom or ville or "—"


def lodging_villes_label(
    rows: list[dict[str, Any]],
    *,
    domicile: str = "",
    jour: int = 0,
    last_jour: int = 0,
) -> str:
    """Villes de nuit depuis Nature = Hébergement ; domicile = départ (J1) / retour (dernier jour)."""
    villes: list[str] = []
    hebergement_rows = [r for r in rows if is_hebergement(r["action"])]
    for row in sorted(hebergement_rows, key=lambda r: r["visite"]):
        ville = lodging_ville(row)
        if not ville:
            continue
        if not villes or villes[-1] != ville:
            villes.append(ville)

    home = normalize_text(domicile)
    if home and villes:
        has_overnight = any(is_overnight_stay(r) for r in hebergement_rows)
        has_checkout = any(not is_overnight_stay(r) for r in hebergement_rows)
        if jour == 1 and has_overnight and villes[0] != home:
            villes.insert(0, home)
        if (
            jour == last_jour
            and has_checkout
            and villes[-1] != home
            and not any(is_overnight_stay(r) and lodging_ville(r) == home for r in hebergement_rows)
        ):
            villes.append(home)

    if len(villes) >= 2:
        return " → ".join(villes)
    if villes:
        return villes[0]
    return "—"


def format_day_villes(summary: dict[str, Any]) -> str:
    return summary.get("lodging_villes_label") or "—"


def day_theme_auto(day_rows: list[dict[str, Any]], summary: dict[str, Any]) -> str:
    activities = [r for r in day_rows if not r["is_trajet_line"]]
    if not activities:
        ville = summary.get("ville") or ""
        return f"{ville} — journée libre" if ville else "Journée libre"

    for row in sorted(activities, key=lambda r: r["visite"]):
        if is_hebergement(row["action"]):
            continue
        if row["nom"]:
            return row["nom"]
    return summary.get("resume") or day_resume(day_rows, limit=2)


def should_highlight_day(summary: dict[str, Any]) -> bool:
    return summary.get("activities", 0) == 0


def phase_jour_bounds(phase: dict[str, Any]) -> tuple[int, int] | None:
    match = re.search(r"Jour\s+(\d+)[–-](\d+)", normalize_text(phase.get("jours")))
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def build_trip_steps(
    phases: list[dict[str, Any]],
    day_summaries: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    by_jour = {summary["jour"]: summary for summary in day_summaries}
    steps: list[dict[str, Any]] = []
    for phase in phases:
        bounds = phase_jour_bounds(phase)
        description = ""
        if bounds:
            from_jour, to_jour = bounds
            themes = [
                normalize_text(by_jour[j]["theme"])
                for j in range(from_jour, to_jour + 1)
                if j in by_jour and normalize_text(by_jour[j]["theme"])
            ]
            if len(themes) == 1:
                description = themes[0]
            elif themes:
                description = f"{themes[0]} … {themes[-1]}"
        steps.append(
            {
                "ville": phase["ville"],
                "dates": format_date_range_fr(phase["start"], phase["end"]),
                "description": description,
            }
        )
    return steps


def fill_empty_day_villes(day_summaries: list[dict[str, Any]]) -> None:
    last_ville = ""
    for summary in day_summaries:
        if summary["ville"]:
            last_ville = summary["ville"]
        elif last_ville:
            summary["ville"] = last_ville
    next_ville = ""
    for summary in reversed(day_summaries):
        if summary["ville"]:
            next_ville = summary["ville"]
        elif next_ville:
            summary["ville"] = next_ville


def max_planned_jour(wb: openpyxl.Workbook, by_jour: dict[int, list[dict[str, Any]]]) -> int:
    from_sheets = 0
    for sheet_name in day_sheets(wb):
        try:
            from_sheets = max(from_sheets, int(sheet_name.split()[1]))
        except (IndexError, ValueError):
            continue
    from_rows = max(by_jour) if by_jour else 0
    return max(from_sheets, from_rows)


def google_maps_url(
    *,
    nom: str = "",
    ville: str = "",
    lat: float | None = None,
    lon: float | None = None,
    existing: str = "",
) -> str | None:
    url = normalize_text(existing)
    lower = url.lower()
    if url and ("google.com/maps" in lower or "maps.google" in lower or "goo.gl/maps" in lower):
        return url
    if lat is not None and lon is not None:
        return f"https://www.google.com/maps/search/?api=1&query={lat},{lon}"
    parts = [normalize_text(part) for part in (nom, ville) if normalize_text(part)]
    if not parts:
        return None
    return f"https://www.google.com/maps/search/?api=1&query={quote(', '.join(parts))}"


def lodging_from_excel(
    rows: list[dict[str, Any]],
    start_date: date,
) -> dict[str, dict[str, Any]]:
    by_ville: dict[str, dict[str, Any]] = {}
    for row in rows:
        if not is_overnight_stay(row):
            continue
        ville = lodging_ville(row)
        if not ville:
            continue
        entry = by_ville.setdefault(
            ville,
            {"ville": ville, "nom": "", "jours": set(), "lat": None, "lon": None, "lien": ""},
        )
        if row["nom"]:
            entry["nom"] = row["nom"]
        if row.get("lat") is not None and row.get("lon") is not None:
            entry["lat"] = row["lat"]
            entry["lon"] = row["lon"]
        if row.get("lien"):
            entry["lien"] = row["lien"]
        elif row.get("site") and not entry["lien"]:
            entry["lien"] = row["site"]
        entry["jours"].add(row["jour"])

    auto_rows: dict[str, dict[str, Any]] = {}
    for ville, entry in by_ville.items():
        jours = sorted(entry["jours"])
        start = date_for_jour(start_date, jours[0])
        end = date_for_jour(start_date, jours[-1])
        nights = len(jours)
        auto_rows[ville] = {
            "ville": ville,
            "nom": entry["nom"] or "À compléter",
            "dates": format_lodging_dates_label(start, end, nights),
            "maps_url": google_maps_url(
                nom=entry["nom"],
                ville=ville,
                lat=entry["lat"],
                lon=entry["lon"],
                existing=entry["lien"],
            ),
            "start": start,
        }
    return auto_rows


def build_lodging_rows(
    rows: list[dict[str, Any]],
    start_date: date,
) -> list[dict[str, Any]]:
    auto_by_ville = lodging_from_excel(rows, start_date)
    return sorted(auto_by_ville.values(), key=lambda row: row["start"])


def resolve_banner_title(config: dict[str, Any], start_date: date) -> str:
    custom = normalize_text(config.get("banner_title"))
    if custom:
        return custom
    month = MOIS_FR[start_date.month]
    if month:
        return f"Vacances {month.capitalize()} {start_date.year}"
    return "Vue d'ensemble"


def primary_ville(rows: list[dict[str, Any]]) -> str:
    counts: Counter[str] = Counter()
    first_order: dict[str, int] = {}
    for row in rows:
        if row["is_trajet_line"] or not row["ville"]:
            continue
        counts[row["ville"]] += 1
        first_order.setdefault(row["ville"], row["visite"])
    if not counts:
        return ""
    return max(counts.keys(), key=lambda v: (counts[v], -first_order[v]))


def villes_for_day(rows: list[dict[str, Any]]) -> list[str]:
    seen: list[str] = []
    for row in sorted(rows, key=lambda r: r["visite"]):
        ville = row["ville"]
        if row["is_trajet_line"] or not ville or ville in seen:
            continue
        seen.append(ville)
    return seen


def day_resume(rows: list[dict[str, Any]], limit: int = 3) -> str:
    visites = [
        row["nom"]
        for row in sorted(rows, key=lambda r: r["visite"])
        if not row["is_trajet_line"] and row["action"].lower() == "visite"
    ]
    if visites:
        head = ", ".join(visites[:limit])
        extra = len(visites) - limit
        return head + (f" (+{extra})" if extra > 0 else "")
    actions = Counter(
        row["action"] or "Activité"
        for row in rows
        if not row["is_trajet_line"]
    )
    if not actions:
        return "Trajets / logistique"
    return ", ".join(f"{n} {label.lower()}" for label, n in actions.most_common(3))


def build_phases_auto(day_summaries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    phases: list[dict[str, Any]] = []
    if not day_summaries:
        return phases

    phase_start_idx = 0
    prev_ville = day_summaries[0]["ville"]
    for idx, summary in enumerate(day_summaries[1:], start=1):
        if summary["ville"] != prev_ville:
            start = day_summaries[phase_start_idx]
            end = day_summaries[idx - 1]
            phases.append(
                {
                    "ville": prev_ville,
                    "start": start["date"],
                    "end": end["date"],
                    "jours": f"Jour {start['jour']}–{end['jour']}",
                    "from_jour": start["jour"],
                    "to_jour": end["jour"],
                }
            )
            phase_start_idx = idx
            prev_ville = summary["ville"]
    start = day_summaries[phase_start_idx]
    end = day_summaries[-1]
    phases.append(
        {
            "ville": prev_ville,
            "start": start["date"],
            "end": end["date"],
            "jours": f"Jour {start['jour']}–{end['jour']}",
            "from_jour": start["jour"],
            "to_jour": end["jour"],
        }
    )
    return phases


def collect_overview_data(
    wb: openpyxl.Workbook,
    config: dict[str, Any],
) -> dict[str, Any]:
    start_date = parse_start_date(config["start_date"])
    resume_limit = int(config.get("day_resume_limit") or 3)

    rows = [row_entry(item) for item in iter_activity_rows(wb)]
    by_jour: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_jour[row["jour"]].append(row)

    planned_last_jour = max_planned_jour(wb, by_jour)
    domicile = normalize_text(config.get("domicile"))

    day_summaries: list[dict[str, Any]] = []
    route_cities: list[str] = []

    for jour in range(1, planned_last_jour + 1):
        day_rows = by_jour.get(jour, [])
        main_ville = primary_ville(day_rows)
        day_villes = villes_for_day(day_rows)
        for ville in day_villes:
            if ville and (not route_cities or route_cities[-1] != ville):
                route_cities.append(ville)

        activities = [r for r in day_rows if not r["is_trajet_line"]]
        visites = [r for r in activities if r["action"].lower() == "visite"]
        prix_rows = [r for r in day_rows if r["prix"] is not None]
        prix_total = sum(r["prix"] for r in prix_rows)
        summary = {
            "jour": jour,
            "date": date_for_jour(start_date, jour),
            "ville": main_ville,
            "villes": day_villes,
            "activities": len(activities),
            "visites": len(visites),
            "prix": prix_total,
            "resume": day_resume(day_rows, limit=resume_limit) if day_rows else "",
            "couleur": jour_color(jour),
        }
        summary["theme"] = day_theme_auto(day_rows, summary)
        summary["lodging_villes_label"] = lodging_villes_label(
            day_rows,
            domicile=domicile,
            jour=jour,
            last_jour=planned_last_jour,
        )
        day_summaries.append(summary)

    fill_empty_day_villes(day_summaries)
    nightly = nightly_lodging_by_jour(rows)
    for summary in day_summaries:
        summary["highlight"] = should_highlight_day(summary)
        summary["nuit"] = nightly.get(summary["jour"])

    jours = [summary["jour"] for summary in day_summaries]
    phases = build_phases_auto(day_summaries)

    by_ville: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"activities": 0, "visites": 0, "prix": 0.0, "jours": set()}
    )
    for row in rows:
        if row["is_trajet_line"] or not row["ville"]:
            continue
        entry = by_ville[row["ville"]]
        entry["activities"] += 1
        entry["jours"].add(row["jour"])
        if row["action"].lower() == "visite":
            entry["visites"] += 1
        if row["prix"] is not None:
            entry["prix"] += row["prix"]

    ville_rows = []
    for ville in sorted(by_ville, key=lambda v: min(by_ville[v]["jours"])):
        entry = by_ville[ville]
        ville_rows.append(
            {
                "ville": ville,
                "activities": entry["activities"],
                "visites": entry["visites"],
                "prix": entry["prix"],
                "jours": len(entry["jours"]),
            }
        )

    prix_total = sum(r["prix"] for r in rows if r["prix"] is not None)
    visites_total = sum(1 for r in rows if r["action"].lower() == "visite")
    activities_total = sum(1 for r in rows if not r["is_trajet_line"])

    if jours:
        period = format_date_range_fr(
            date_for_jour(start_date, jours[0]),
            date_for_jour(start_date, jours[-1]),
        )
    else:
        period = ""

    route = " → ".join(route_cities)
    trip_steps = build_trip_steps(phases, day_summaries)
    lodging_rows = build_lodging_rows(rows, start_date)

    return {
        "banner_title": resolve_banner_title(config, start_date),
        "generated_at": datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC"),
        "period": period,
        "route": route,
        "jours_count": len(jours),
        "activities_total": activities_total,
        "visites_total": visites_total,
        "prix_total": prix_total,
        "day_summaries": day_summaries,
        "phases": phases,
        "trip_steps": trip_steps,
        "lodging_rows": lodging_rows,
        "ville_rows": ville_rows,
    }


def json_default(value: Any) -> Any:
    if isinstance(value, date):
        return value.isoformat()
    raise TypeError(f"Type non serialisable: {type(value)!r}")


def write_overview_snapshot(data: dict[str, Any], config: dict[str, Any]) -> Path:
    snapshot = {
        "generated_at": data["generated_at"],
        "config": {
            "start_date": config["start_date"],
            "sheet_name": config["sheet_name"],
            "domicile": config.get("domicile"),
        },
        "summary": {
            "period": data["period"],
            "route": data["route"],
            "jours_count": data["jours_count"],
            "activities_total": data["activities_total"],
            "visites_total": data["visites_total"],
            "prix_total": data["prix_total"],
        },
        "phases": data["phases"],
        "steps": data["trip_steps"],
        "lodging": data["lodging_rows"],
        "by_day": data["day_summaries"],
        "by_ville": data["ville_rows"],
    }
    path = data_dir() / "overview.json"
    path.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2, default=json_default),
        encoding="utf-8",
    )
    return path


def set_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: Any,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def set_link_cell(
    ws: Worksheet,
    row: int,
    col: int,
    url: str,
    label: str = "Google Maps",
    *,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    cell = ws.cell(row=row, column=col, value=label)
    cell.hyperlink = url
    cell.font = LINK_FONT
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def merge_row(ws: Worksheet, row: int, col_start: int, col_end: int) -> None:
    if col_end > col_start:
        ws.merge_cells(
            start_row=row,
            start_column=col_start,
            end_row=row,
            end_column=col_end,
        )


def write_table_header(ws: Worksheet, row: int, headers: list[str]) -> None:
    ws.row_dimensions[row].height = 24
    for col, header in enumerate(headers, start=1):
        set_cell(
            ws,
            row,
            col,
            header,
            font=TABLE_HEADER_FONT,
            fill=TABLE_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=HEADER_BORDER,
        )


def apply_fixed_column_widths(ws: Worksheet) -> None:
    for col, width in enumerate(OVERVIEW_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def clear_sheet(ws: Worksheet) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def ensure_overview_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Worksheet:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name, 0)


def write_day_row(
    ws: Worksheet,
    row: int,
    summary: dict[str, Any],
) -> None:
    ws.row_dimensions[row].height = DATA_ROW_HEIGHT
    jour = summary["jour"]
    day_color = summary.get("couleur") or jour_color(jour)
    row_fill = HIGHLIGHT_FILL if summary.get("highlight") else fill_from_hex(day_color, light=True)
    jour_fill = fill_from_hex(day_color)
    night_text = format_night_cell(summary.get("nuit"))

    values = [
        jour,
        format_date_fr(summary["date"]),
        format_day_villes(summary),
        night_text,
        summary["activities"],
        format_budget_cell(summary["prix"], summary["activities"]),
    ]
    for col, value in enumerate(values, start=1):
        if col == 1:
            set_cell(
                ws,
                row,
                col,
                value,
                font=JOUR_FONT,
                fill=jour_fill,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=TABLE_BORDER,
            )
            continue

        font = LODGING_FONT if col == 4 and night_text != "—" else BODY_FONT
        cell_fill = row_fill
        if col == 4 and isinstance(value, str) and is_placeholder(value):
            cell_fill = HIGHLIGHT_FILL

        alignment = Alignment(vertical="center")
        if col in (1, 5, 6):
            alignment = Alignment(horizontal="center", vertical="center")
        elif col == 2:
            alignment = Alignment(horizontal="center", vertical="center")
        elif col == 3:
            alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        set_cell(
            ws,
            row,
            col,
            value,
            font=font,
            fill=cell_fill,
            alignment=alignment,
            border=TABLE_BORDER,
        )


def write_section_title(ws: Worksheet, row: int, title: str) -> None:
    set_cell(
        ws,
        row,
        1,
        title,
        font=SECTION_HEADER_FONT,
        fill=SECTION_HEADER_FILL,
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    merge_row(ws, row, 1, OVERVIEW_MAX_COL)
    ws.row_dimensions[row].height = 22


def write_lodging_table(ws: Worksheet, start_row: int, lodging_rows: list[dict[str, Any]]) -> int:
    row = start_row
    ws.row_dimensions[row].height = 24
    lodging_headers = [
        (1, "Ville", False),
        (2, "Période", False),
        (3, "Hébergement", True),
        (6, "Lien", False),
    ]
    for col, label, merge_rest in lodging_headers:
        set_cell(
            ws, row, col, label,
            font=TABLE_HEADER_FONT,
            fill=TABLE_HEADER_FILL,
            alignment=Alignment(horizontal="center", vertical="center"),
            border=HEADER_BORDER,
        )
        if merge_rest:
            merge_row(ws, row, col, OVERVIEW_MAX_COL - 1)
            for extra_col in range(col + 1, OVERVIEW_MAX_COL):
                cell = ws.cell(row=row, column=extra_col)
                cell.fill = TABLE_HEADER_FILL
                cell.border = HEADER_BORDER
    row += 1

    if not lodging_rows:
        set_cell(ws, row, 1, "Aucun hébergement renseigné", font=BODY_FONT)
        merge_row(ws, row, 1, OVERVIEW_MAX_COL)
        return row + 1

    for entry in lodging_rows:
        ws.row_dimensions[row].height = DATA_ROW_HEIGHT
        set_cell(
            ws, row, 1, entry["ville"],
            font=BODY_FONT, fill=HIGHLIGHT_FILL,
            alignment=Alignment(horizontal="center", vertical="center"),
            border=TABLE_BORDER,
        )
        set_cell(
            ws, row, 2, entry["dates"],
            font=BODY_FONT,
            alignment=Alignment(horizontal="center", vertical="center"),
            border=TABLE_BORDER,
        )
        set_cell(
            ws, row, 3, entry.get("nom") or "À compléter",
            font=LODGING_FONT,
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=TABLE_BORDER,
        )
        merge_row(ws, row, 3, OVERVIEW_MAX_COL - 1)
        for col in range(4, OVERVIEW_MAX_COL):
            cell = ws.cell(row=row, column=col)
            cell.border = TABLE_BORDER
        maps_url = entry.get("maps_url") or ""
        if maps_url:
            set_link_cell(
                ws, row, OVERVIEW_MAX_COL, maps_url, label="Carte",
                alignment=Alignment(horizontal="center", vertical="center"),
                border=TABLE_BORDER,
            )
        else:
            set_cell(
                ws, row, OVERVIEW_MAX_COL, "—",
                font=BODY_FONT,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=TABLE_BORDER,
            )
        row += 1
    return row


def write_totals_row(ws: Worksheet, row: int, data: dict[str, Any]) -> None:
    ws.row_dimensions[row].height = 22
    label = (
        f"{data['jours_count']} jour(s) · "
        f"{data['activities_total']} activité(s) · "
        f"{data['visites_total']} visite(s)"
    )
    prix = data.get("prix_total") or 0
    if prix:
        label += f" · budget estimé ~{prix:g} €".replace(".", ",")
    set_cell(
        ws, row, 1, label,
        font=TOTAL_FONT,
        alignment=Alignment(horizontal="left", vertical="center"),
    )
    merge_row(ws, row, 1, OVERVIEW_MAX_COL)


def write_auto_generated_notice(ws: Worksheet, row: int) -> int:
    set_cell(
        ws,
        row,
        1,
        "Feuille générée automatiquement — toute modification manuelle sera remplacée à la prochaine génération.",
        font=AUTO_GENERATED_FONT,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    merge_row(ws, row, 1, OVERVIEW_MAX_COL)
    ws.row_dimensions[row].height = 16
    return row + 1


def render_overview_sheet(ws: Worksheet, data: dict[str, Any]) -> None:
    clear_sheet(ws)
    row = 1

    set_cell(
        ws, row, 1, data["banner_title"],
        font=BANNER_FONT, fill=BANNER_FILL,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    merge_row(ws, row, 1, OVERVIEW_MAX_COL)
    ws.row_dimensions[row].height = 28
    row += 1

    if data.get("period"):
        set_cell(
            ws, row, 1, data["period"],
            font=SUBTITLE_FONT,
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        merge_row(ws, row, 1, OVERVIEW_MAX_COL)
        row += 1

    row = write_auto_generated_notice(ws, row)
    row += 1
    write_table_header(
        ws,
        row,
        ["Jour", "Date", "Villes du jour", "Nuit à", "Activités", "Budget"],
    )
    header_row = row
    row += 1

    for summary in data.get("day_summaries") or []:
        write_day_row(ws, row, summary)
        row += 1

    row += 1
    write_section_title(ws, row, "Hébergements du voyage")
    row += 1
    row = write_lodging_table(ws, row, data.get("lodging_rows") or [])

    row += 1
    write_totals_row(ws, row, data)

    ws.freeze_panes = f"A{header_row + 1}"
    apply_fixed_column_widths(ws)


def write_snapshot_from_excel(excel_path: Path, config: dict[str, Any]) -> Path | None:
    """Ecrit data/overview.json sans modifier la feuille Excel."""
    if not config.get("write_snapshot", True):
        return None
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    data = collect_overview_data(wb, config)
    return write_overview_snapshot(data, config)


def run_build(
    excel_path: Path,
    config: dict[str, Any],
    *,
    config_path: Path | None = None,
    dry_run: bool = False,
    snapshot_only: bool = False,
) -> None:
    if snapshot_only:
        if dry_run:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            data = collect_overview_data(wb, config)
            print(f"(dry-run) Snapshot overview.json")
            print(f"  Période : {data['period']}")
            print(f"  {len(data['day_summaries'])} jour(s), {data['activities_total']} activité(s)")
            return
        snapshot_path = write_snapshot_from_excel(excel_path, config)
        if snapshot_path:
            print(f"Snapshot overview : {snapshot_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    data = collect_overview_data(wb, config)
    sheet_name = config["sheet_name"]

    if dry_run:
        print(f"(dry-run) Vue d'ensemble pour {excel_path.name}")
        print(f"  Config : {config_path or default_overview_config_path()}")
        print(f"  Période : {data['period']}")
        print(f"  Itinéraire : {data['route']}")
        print(f"  {len(data['day_summaries'])} jour(s), {data['activities_total']} activité(s)")
        return

    ws = ensure_overview_sheet(wb, sheet_name)
    render_overview_sheet(ws, data)

    backup_path = backup_excel(excel_path)
    wb.save(excel_path)

    snapshot_path = None
    if config.get("write_snapshot", True):
        snapshot_path = write_overview_snapshot(data, config)

    print(f"Vue d'ensemble regeneree : {excel_path}")
    print(f"  Feuille : {sheet_name}")
    print(f"  Periode : {data['period']}")
    print(f"  Backup : {backup_path}")
    if snapshot_path:
        print(f"  Snapshot : {snapshot_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Regenere la feuille Vue d'ensemble depuis les feuilles Jour N."
    )
    parser.add_argument("excel", nargs="?", default=str(default_excel_path()))
    parser.add_argument(
        "--config",
        default=str(default_overview_config_path()),
        help="Fichier JSON de configuration (defaut: data/overview_config.json).",
    )
    parser.add_argument(
        "--start-date",
        default=None,
        help="Surcharge la date du jour 1 (AAAA-MM-JJ).",
    )
    parser.add_argument("--dry-run", action="store_true", help="Afficher sans ecrire dans Excel")
    parser.add_argument(
        "--snapshot-only",
        action="store_true",
        help="Ecrire uniquement data/overview.json (sans modifier Excel).",
    )
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    if not excel_path.exists():
        raise SystemExit(f"Fichier introuvable: {excel_path}")

    config_path = Path(args.config).resolve()
    try:
        config = resolve_overview_config(
            config_path,
            start_date=args.start_date,
        )
        parse_start_date(config["start_date"])
    except (ValueError, json.JSONDecodeError) as exc:
        raise SystemExit(str(exc)) from exc

    run_build(
        excel_path,
        config,
        config_path=config_path,
        dry_run=args.dry_run,
        snapshot_only=args.snapshot_only,
    )


if __name__ == "__main__":
    main()
