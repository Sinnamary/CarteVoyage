"""Verification du classeur de planning."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from outils.cli_args import VerifyWorkbookArgs
from outils.excel_utils import (
    LINKS_SHEET,
    OVERVIEW_SHEET,
    PLANNING_COLUMNS,
    build_listes_ranges,
    day_sheets,
    default_excel_path,
    find_duplicate_ordre_labels,
    find_ordre_collisions,
    iter_activity_rows,
    normalize_text,
    row_to_point,
    sync_listes_validations,
)
from outils.overview_config import load_overview_config

MAP_COLUMNS = ("Latitude", "Longitude")
EXPECTED_DAY_COUNT = 12


def check_ordre_text_format(excel_path: Path, wb: openpyxl.Workbook) -> list[str]:
    issues: list[str] = []
    wb_fmt = openpyxl.load_workbook(excel_path, data_only=False)
    for sheet_name in day_sheets(wb):
        ws = wb_fmt[sheet_name]
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(r, 1)
            etape = cell.value
            if etape is None:
                continue
            text = str(etape)
            if ".10" in text and cell.number_format != "@":
                issues.append(
                    f"{sheet_name} L{r}: {text!r} doit être en texte (@), "
                    + f"format={cell.number_format!r}"
                )
    return issues


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Verifie la structure du classeur Excel de voyage."
    )
    _ = parser.add_argument("excel", nargs="?", default=str(default_excel_path()))
    args = parser.parse_args(namespace=VerifyWorkbookArgs())

    path = Path(args.excel).resolve()
    if not path.exists():
        raise SystemExit(f"Fichier introuvable: {path}")

    overview_config = load_overview_config()
    overview_sheet = overview_config.get("sheet_name") or OVERVIEW_SHEET
    verify_markers = overview_config.get("verify_markers") or []

    wb = openpyxl.load_workbook(path, data_only=True)
    errors: list[str] = []
    warnings: list[str] = []

    day_sheet_names = [f"Jour {d}" for d in range(1, EXPECTED_DAY_COUNT + 1)]
    required_sheets = [overview_sheet, LINKS_SHEET, "Listes", *day_sheet_names]
    missing_sheets = [name for name in required_sheets if name not in wb.sheetnames]
    extra_sheets = [name for name in wb.sheetnames if name not in required_sheets]
    if missing_sheets:
        errors.append(f"Feuilles manquantes: {missing_sheets}")
    if extra_sheets:
        errors.append(f"Feuilles inattendues: {extra_sheets}")
    if LINKS_SHEET in wb.sheetnames and "Listes" in wb.sheetnames:
        links_idx = wb.sheetnames.index(LINKS_SHEET)
        listes_idx = wb.sheetnames.index("Listes")
        if links_idx != listes_idx - 1:
            errors.append(
                f"La feuille {LINKS_SHEET!r} doit etre placee juste avant Listes "
                + f"(ordre actuel: {wb.sheetnames})"
            )

    listes = wb["Listes"]
    if listes.sheet_state != "hidden":
        warnings.append("Feuille Listes devrait être masquée")

    wb_validations = openpyxl.load_workbook(path)
    expected_ranges = build_listes_ranges(wb_validations)
    pending = sync_listes_validations(wb_validations)
    if pending:
        warnings.append(
            "Listes deroulantes desynchronisees: lancer preparer_excel.py "
            + "ou scripts/outils/sync_listes_validations.py "
            + f"(ex. Ville attendue {expected_ranges.get('E', '?')})"
        )

    if overview_sheet not in wb.sheetnames:
        errors.append(f"Feuille manquante: {overview_sheet}")
    elif verify_markers:
        overview = wb[overview_sheet]
        overview_text = " ".join(
            str(cell.value) for row in overview.iter_rows() for cell in row if cell.value
        )
        warnings.extend(
            f"{overview_sheet}: '{marker}' absent"
            for marker in verify_markers
            if marker not in overview_text
        )

    for sheet_name in day_sheets(wb):
        ws = wb[sheet_name]
        header_names = {
            normalize_text(ws.cell(2, c).value)
            for c in range(1, ws.max_column + 1)
            if ws.cell(2, c).value
        }
        missing = [col for col in PLANNING_COLUMNS.values() if col not in header_names]
        if missing:
            errors.append(f"{sheet_name}: colonnes manquantes {missing}")

        missing_map = [col for col in MAP_COLUMNS if col not in header_names]
        if missing_map:
            warnings.append(
                f"{sheet_name}: {', '.join(missing_map)} absentes (ajoutées par geocode_excel.py)"
            )

        banner = normalize_text(ws.cell(1, 1).value)
        if not banner.startswith(sheet_name):
            warnings.append(f"{sheet_name}: bannière inattendue ({banner[:50]}...)")

    for (jour, visite), noms in sorted(find_ordre_collisions(wb).items()):
        warnings.append(f"Collision ordre {jour}.{visite}: {', '.join(noms)}")

    for label, locs in sorted(find_duplicate_ordre_labels(wb).items()):
        warnings.append(f"N° étape {label} répété: {', '.join(locs)}")

    errors.extend(check_ordre_text_format(path, wb))

    items = list(iter_activity_rows(wb))
    with_coords = sum(1 for item in items if row_to_point(item))

    print(f"Fichier: {path}")
    print(f"Lignes activité: {len(items)}")
    print(f"Avec coordonnées (carte): {with_coords}")
    print(f"Sans coordonnées: {len(items) - with_coords}")

    if warnings:
        print("\nAvertissements:")
        for w in warnings:
            print(f"  - {w}")

    if errors:
        print("\nProblèmes:")
        for e in errors:
            print(f"  - {e}")
        raise SystemExit(1)

    print("\nOK: classeur valide")


if __name__ == "__main__":
    main()
