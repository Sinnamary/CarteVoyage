"""Utilitaires pour lire/écrire le classeur de planning (feuilles Jour 1, Jour 2, …)."""

from __future__ import annotations

import re
import shutil
from collections.abc import Iterator
from datetime import datetime
from pathlib import Path
from typing import Any, TypedDict

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_EXCEL_NAME = "Voyage Aout 2026.xlsx"

OVERVIEW_SHEET = "Vue d'ensemble"
# Ancien nom sans apostrophe (fichiers Google Drive / Excel en ligne).
OVERVIEW_SHEET_LEGACY = "Vue densemble"
LINKS_SHEET = "Liens"
IGNORE_SHEETS = {OVERVIEW_SHEET, LINKS_SHEET, "Listes"}
DAY_HEADER_ROW = 2
DATA_START_ROW = 3

PLANNING_COLUMNS = {
    "Ordre": "N° étape",
    "Nom": "Lieu",
    "Action": "Nature",
    "Type": "Catégorie",
    "Remarque": "Quartier",
    "Ville": "Ville",
    "Billet": "Réservation",
    "Prix": "Prix (€)",
    "Ouverture": "Heure début",
    "Fermeture": "Heure fin",
    "Site": "Site web",
}

MAP_EXTRA_COLUMNS = ["Latitude", "Longitude", "Lien"]
OPTIONAL_COLUMNS = ["City Card"]
SKIP_LIEU = {"", "Journée à planifier"}

DAY_COLORS = [
    "#e74c3c",
    "#27ae60",
    "#3498db",
    "#9b59b6",
    "#e67e22",
    "#1abc9c",
    "#f39c12",
    "#2c3e50",
    "#d35400",
    "#16a085",
]

ORDRE_RE = re.compile(r"^\s*(\d+)[.,](\d+)\s*$")


class ActivityRow(TypedDict):
    row_idx: int
    jour: int
    visite: int
    ordre: int
    ordre_label: str
    nom: str
    col_index: dict[str, int]
    row: tuple[Any, ...]
    ws: Worksheet
    header_row_idx: int
    sheet_name: str


def project_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent


def excel_dir() -> Path:
    path = project_root() / "excel"
    path.mkdir(exist_ok=True)
    return path


def data_dir() -> Path:
    path = project_root() / "data"
    path.mkdir(exist_ok=True)
    return path


def web_dir() -> Path:
    path = project_root() / "web"
    path.mkdir(exist_ok=True)
    return path


def default_excel_path() -> Path:
    return excel_dir() / DEFAULT_EXCEL_NAME


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_trajet_line(nom: str) -> bool:
    lower = normalize_text(nom).lower()
    return lower.startswith(("trajet ", "retour "))


def is_lodging_action(action: str) -> bool:
    return normalize_text(action).lower() in ("hébergement", "hebergement")


def lodging_ville(nom: str, ville: str) -> str:
    """Ville affichée pour l'hébergement (le nom prime pour les cas hors colonne Ville)."""
    if "ennevelin" in normalize_text(nom).lower():
        return "Ennevelin"
    return normalize_text(ville)


def lodging_stays_for_day(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Lignes Hébergement d'un jour, triées par N° étape (visite croissante)."""
    stays = [r for r in rows if is_lodging_action(r.get("action", ""))]
    stays.sort(key=lambda r: r["visite"])
    return stays


def lodging_evening_stay(
    stays: list[dict[str, Any]],
    *,
    jour: int,
    last_jour: int,
) -> dict[str, Any] | None:
    """Point de nuitée : dernière ligne Hébergement du jour (ordre de visite).

    Dernier jour avec une seule ligne = check-out matinal, pas de nuit sur place.
    """
    if not stays:
        return None
    if jour == last_jour and len(stays) == 1:
        return None
    return stays[-1]


def lodging_villes_label(
    rows: list[dict[str, Any]],
    *,
    domicile: str = "",
    jour: int = 0,
    last_jour: int = 0,
) -> str:
    """Villes du jour : 1re ligne Hébergement = départ, dernière = arrivée (N° étape)."""
    stays = lodging_stays_for_day(rows)
    if not stays:
        return "—"

    morning_v = lodging_ville(stays[0]["nom"], stays[0]["ville"])
    evening_v = lodging_ville(stays[-1]["nom"], stays[-1]["ville"])

    villes: list[str] = []
    if morning_v:
        villes.append(morning_v)
    if evening_v and evening_v != morning_v:
        villes.append(evening_v)

    home = normalize_text(domicile)
    if home:
        has_evening = len(stays) >= 2 or (len(stays) == 1 and jour != last_jour)
        has_morning = len(stays) >= 2 or (len(stays) == 1 and jour == last_jour)
        if jour == 1 and has_evening and villes and villes[0] != home:
            villes.insert(0, home)
        if (
            jour == last_jour
            and has_morning
            and villes
            and villes[-1] != home
            and evening_v != home
        ):
            villes.append(home)

    if len(villes) >= 2:
        return " → ".join(villes)
    if villes:
        return villes[0]
    return "—"


LODGING_ROLE_LABELS = {
    "depart_matin": "Départ matin",
    "arrivee_soir": "Arrivée soir",
    "check_out": "Check-out (matin)",
    "check_in": "Arrivée / nuit",
}


def _lodging_line_role(*, index: int, count: int, jour: int, last_jour: int) -> str:
    if count >= 2:
        if index == 0:
            return "depart_matin"
        if index == count - 1:
            return "arrivee_soir"
    if jour == last_jour:
        return "check_out"
    return "check_in"


def activity_rows_from_workbook(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """Lignes activité du classeur (jour, visite, ordre, nom, ville, action)."""
    rows: list[dict[str, Any]] = []
    for item in iter_activity_rows(wb):
        ci = item["col_index"]
        row = item["row"]
        rows.append(
            {
                "jour": item["jour"],
                "visite": item["visite"],
                "ordre": item["ordre_label"],
                "nom": item["nom"],
                "ville": normalize_text(cell_value(row, ci, "Ville")),
                "action": normalize_text(cell_value(row, ci, "Action")),
            }
        )
    return rows


def build_lodging_audit(
    rows: list[dict[str, Any]],
    *,
    domicile: str = "",
    jours: list[int] | None = None,
    overview_by_day: dict[int, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Trace détaillée de la détermination des hébergements par jour."""
    if not jours:
        jours = sorted({int(r["jour"]) for r in rows})
    last_jour = max(jours) if jours else 0
    home = normalize_text(domicile)

    by_jour: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        by_jour.setdefault(int(row["jour"]), []).append(row)

    rule = (
        "Les lignes Nature = Hébergement sont triées par N° étape (visite croissante). "
        "La première = point de départ le matin, la dernière = arrivée le soir et nuitée. "
        "Au dernier jour, une seule ligne = check-out matinal (pas de nuit sur place). "
    )
    if home:
        rule += (
            f"Le domicile « {home} » est ajouté au libellé « Villes du jour » "
            "le jour 1 (départ depuis le domicile) et le dernier jour (retour)."
        )

    days_out: list[dict[str, Any]] = []
    for jour in jours:
        day_rows = sorted(by_jour.get(jour, []), key=lambda r: r["visite"])
        stays = lodging_stays_for_day(day_rows)
        flags: list[str] = []
        if jour == 1:
            flags.append("premier_jour")
        if jour == last_jour:
            flags.append("dernier_jour")

        first_activity = day_rows[0] if day_rows else None
        lodging_lines: list[dict[str, Any]] = []
        for index, stay in enumerate(stays):
            role = _lodging_line_role(index=index, count=len(stays), jour=jour, last_jour=last_jour)
            lodging_lines.append(
                {
                    "ordre": stay.get("ordre") or f"{stay['jour']}.{stay['visite']}",
                    "visite": stay["visite"],
                    "nom": stay["nom"],
                    "ville": stay.get("ville", ""),
                    "ville_display": lodging_ville(stay["nom"], stay.get("ville", "")),
                    "role": role,
                    "role_label": LODGING_ROLE_LABELS.get(role, role),
                }
            )

        evening_row = lodging_evening_stay(stays, jour=jour, last_jour=last_jour)
        night: dict[str, Any] | None = None
        if evening_row:
            night = {
                "ordre": evening_row.get("ordre")
                or f"{evening_row['jour']}.{evening_row['visite']}",
                "visite": evening_row["visite"],
                "nom": evening_row["nom"],
                "ville": lodging_ville(evening_row["nom"], evening_row.get("ville", "")),
            }
            night_reason = (
                f"Dernière ligne Hébergement : {night['ordre']} "
                f"(visite {night['visite']}) → nuit à {night['ville']}"
            )
        elif stays and jour == last_jour and len(stays) == 1:
            night_reason = (
                f"Une seule ligne Hébergement ({lodging_lines[0]['ordre']}) "
                "au dernier jour = check-out matinal, pas de nuit sur place"
            )
        elif not stays:
            night_reason = "Aucune ligne Nature = Hébergement ce jour"
        else:
            night_reason = "Pas de nuitée retenue"

        villes_label = lodging_villes_label(
            day_rows, domicile=domicile, jour=jour, last_jour=last_jour
        )
        steps = [
            f"{line['role_label']} : {line['nom']} ({line['ordre']})" for line in lodging_lines
        ]
        if home and jour == 1 and stays:
            evening_v = lodging_ville(stays[-1]["nom"], stays[-1]["ville"])
            has_evening = len(stays) >= 2 or (len(stays) == 1 and jour != last_jour)
            if has_evening and evening_v and villes_label.startswith(home):
                steps.append(f"Domicile « {home} » ajouté en tête (jour 1, nuit hors domicile)")
        if home and jour == last_jour and stays:
            has_morning = len(stays) >= 2 or len(stays) == 1
            evening_v = lodging_ville(stays[-1]["nom"], stays[-1]["ville"])
            if has_morning and villes_label.endswith(home) and evening_v != home:
                steps.append(f"Domicile « {home} » ajouté en fin (dernier jour, retour)")

        notes: list[str] = []
        if jour == 1 and first_activity and not is_lodging_action(first_activity.get("action", "")):
            notes.append(
                f"La 1re activité ({first_activity.get('ordre')} {first_activity.get('nom')}, "
                f"{first_activity.get('action') or '—'}) n'est pas un hébergement : "
                + "le voyage précède l'enregistrement à l'hôtel."
            )
        if jour == last_jour and stays and len(stays) == 1:
            notes.append(
                "Dernier jour : une seule ligne Hébergement = départ matin et retour au domicile."
            )

        overview_day = (overview_by_day or {}).get(jour) or {}
        overview_nuit = overview_day.get("nuit")
        overview_villes = overview_day.get("lodging_villes_label")
        overview_nuit_text = "—"
        if overview_nuit:
            nom = overview_nuit.get("nom") or ""
            ville = overview_nuit.get("ville") or ""
            overview_nuit_text = (
                f"{nom} ({ville})" if ville and ville.lower() not in nom.lower() else (nom or ville)
            )

        match_nuit = None
        if overview_nuit is not None:
            if night and overview_nuit:
                match_nuit = normalize_text(night.get("nom")) == normalize_text(
                    overview_nuit.get("nom")
                ) and normalize_text(night.get("ville")) == normalize_text(
                    overview_nuit.get("ville")
                )
            else:
                match_nuit = night is None and overview_nuit is None

        days_out.append(
            {
                "jour": jour,
                "flags": flags,
                "first_activity": (
                    {
                        "ordre": first_activity.get("ordre"),
                        "nom": first_activity.get("nom"),
                        "action": first_activity.get("action"),
                    }
                    if first_activity
                    else None
                ),
                "lodging_lines": lodging_lines,
                "night": night,
                "night_reason": night_reason,
                "villes_label": villes_label,
                "steps": steps,
                "notes": notes,
                "overview_nuit": overview_nuit_text if overview_nuit else None,
                "overview_villes_label": overview_villes,
                "match_overview_nuit": match_nuit,
            }
        )

    return {
        "rule": rule,
        "domicile": home or None,
        "last_jour": last_jour,
        "days": days_out,
    }


def parse_ordre(value: Any) -> dict[str, int] | None:
    if value is None or value == "":
        return None

    text = normalize_text(value).replace(",", ".")
    if not text:
        return None

    match = ORDRE_RE.match(text)
    if not match and isinstance(value, (int, float)):
        text = f"{float(value):g}".replace(",", ".")
        match = ORDRE_RE.match(text)

    if not match:
        return None

    return {"jour": int(match.group(1)), "visite": int(match.group(2))}


def jour_color(jour: int) -> str:
    return DAY_COLORS[(jour - 1) % len(DAY_COLORS)]


def is_day_sheet(sheet_name: str) -> bool:
    return sheet_name.startswith("Jour ") and sheet_name not in IGNORE_SHEETS


def day_sheets(wb: openpyxl.Workbook) -> list[str]:
    return [name for name in wb.sheetnames if is_day_sheet(name)]


def build_planning_col_index(header_row: tuple[Any, ...]) -> dict[str, int]:
    index: dict[str, int] = {}
    excel_to_internal = {excel: internal for internal, excel in PLANNING_COLUMNS.items()}

    for col_idx, value in enumerate(header_row):
        name = normalize_text(value)
        if not name:
            continue
        if name in excel_to_internal:
            internal = excel_to_internal[name]
            if internal not in index:
                index[internal] = col_idx
        elif name in MAP_EXTRA_COLUMNS + OPTIONAL_COLUMNS and name not in index:
            index[name] = col_idx

    return index


def ensure_map_columns(ws: Worksheet) -> dict[str, int]:
    header = [cell.value for cell in ws[DAY_HEADER_ROW]]
    col_index = build_planning_col_index(tuple(header))

    next_col = len(header) + 1
    while next_col > 1 and header[next_col - 2] is None:
        next_col -= 1
    if next_col <= len(header):
        next_col = len(header) + 1

    for col_name in MAP_EXTRA_COLUMNS:
        if col_name not in col_index:
            ws.cell(row=DAY_HEADER_ROW, column=next_col, value=col_name)
            col_index[col_name] = next_col - 1
            next_col += 1

    return col_index


def cell_value(row: tuple[Any, ...], col_index: dict[str, int], column: str) -> Any:
    if column not in col_index:
        return None
    idx = col_index[column]
    if idx >= len(row):
        return None
    return row[idx]


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_prix(value: Any) -> float | None:
    return parse_float(value)


def has_coordinates(row: tuple[Any, ...], col_index: dict[str, int]) -> bool:
    lat = parse_float(cell_value(row, col_index, "Latitude"))
    lon = parse_float(cell_value(row, col_index, "Longitude"))
    return lat is not None and lon is not None


def backup_excel(excel_path: Path) -> Path:
    backup_dir = excel_path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    backup_path = backup_dir / (excel_path.stem + ".backup" + excel_path.suffix)
    shutil.copy2(excel_path, backup_path)
    return backup_path


def backup_excel_timestamped(excel_path: Path) -> Path | None:
    """Copie horodatee avant remplacement (ex. telechargement Drive)."""
    if not excel_path.exists():
        return None

    backup_dir = excel_path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_dir / f"{excel_path.stem}.backup.{stamp}{excel_path.suffix}"
    shutil.copy2(excel_path, backup_path)
    return backup_path


LISTES_SHEET = "Listes"
LISTES_DATA_START_ROW = 2
# Colonnes de la feuille Listes -> plages utilisees par les listes deroulantes.
LISTES_COLUMN_LETTERS = ("A", "B", "C", "D", "E")


def listes_end_row(ws: Worksheet, col_idx: int) -> int:
    """Derniere ligne non vide d'une colonne de listes (a partir de la ligne 2)."""
    end_row = LISTES_DATA_START_ROW - 1
    for row_idx in range(LISTES_DATA_START_ROW, ws.max_row + 1):
        if normalize_text(ws.cell(row_idx, col_idx).value):
            end_row = row_idx
    return end_row


def listes_range_formula(col_letter: str, end_row: int) -> str:
    return f"Listes!${col_letter}${LISTES_DATA_START_ROW}:${col_letter}${end_row}"


def build_listes_ranges(wb: openpyxl.Workbook) -> dict[str, str]:
    """Calcule les plages Listes!$X$2:$X$N pour chaque colonne de listes."""
    ws = wb[LISTES_SHEET]
    ranges: dict[str, str] = {}
    for col_idx, col_letter in enumerate(LISTES_COLUMN_LETTERS, start=1):
        end_row = listes_end_row(ws, col_idx)
        if end_row >= LISTES_DATA_START_ROW:
            ranges[col_letter] = listes_range_formula(col_letter, end_row)
    return ranges


def sync_listes_validations(wb: openpyxl.Workbook) -> list[str]:
    """
    Met a jour les validations des feuilles Jour pour couvrir toute la feuille Listes.
    Retourne la liste des changements effectues.
    """
    ranges = build_listes_ranges(wb)
    changes: list[str] = []

    for sheet_name in day_sheets(wb):
        ws = wb[sheet_name]
        for dv in ws.data_validations.dataValidation:
            formula = str(dv.formula1 or "")
            if not formula.startswith("Listes!$"):
                continue
            for col_letter, new_range in ranges.items():
                prefix = f"Listes!${col_letter}$"
                if not formula.startswith(prefix):
                    continue
                if formula != new_range:
                    dv.formula1 = new_range
                    changes.append(f"{sheet_name}: {formula} -> {new_range}")
                break

    return changes


def ville_for_row(row: tuple[Any, ...], col_index: dict[str, int]) -> str:
    return normalize_text(cell_value(row, col_index, "Ville"))


def iter_activity_rows(wb: openpyxl.Workbook) -> Iterator[ActivityRow]:
    for sheet_name in day_sheets(wb):
        ws = wb[sheet_name]
        header = next(
            ws.iter_rows(min_row=DAY_HEADER_ROW, max_row=DAY_HEADER_ROW, values_only=True)
        )
        col_index = build_planning_col_index(header)
        if "Nom" not in col_index or "Ordre" not in col_index:
            continue

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            nom = normalize_text(cell_value(row, col_index, "Nom"))
            if nom in SKIP_LIEU:
                continue

            ordre_raw = cell_value(row, col_index, "Ordre")
            parsed = parse_ordre(ordre_raw)
            if not parsed:
                continue

            ordre_label = normalize_text(ordre_raw).replace(",", ".")
            if not ordre_label:
                ordre_label = f"{parsed['jour']}.{parsed['visite']}"

            yield {
                "row_idx": row_idx,
                "jour": parsed["jour"],
                "visite": parsed["visite"],
                "ordre": parsed["visite"],
                "ordre_label": ordre_label,
                "nom": nom,
                "col_index": col_index,
                "row": row,
                "ws": ws,
                "header_row_idx": DAY_HEADER_ROW,
                "sheet_name": sheet_name,
            }


def row_to_point(item: ActivityRow) -> dict[str, Any] | None:
    row = item["row"]
    col_index = item["col_index"]

    lat = parse_float(cell_value(row, col_index, "Latitude"))
    lon = parse_float(cell_value(row, col_index, "Longitude"))
    if lat is None or lon is None:
        return None

    lien = normalize_text(cell_value(row, col_index, "Lien"))
    site = normalize_text(cell_value(row, col_index, "Site"))
    url = lien or site or None

    def field(name: str) -> Any:
        value = cell_value(row, col_index, name)
        if value is None or value == "":
            return None
        return value

    jour = item["jour"]
    visite = item["visite"]
    return {
        "id": f"{jour}.{visite}-{item['row_idx']}",
        "ordre": item["ordre"],
        "ordre_label": item["ordre_label"],
        "jour": jour,
        "visite": visite,
        "nom": item["nom"],
        "ville": field("Ville"),
        "lat": lat,
        "lon": lon,
        "lien": url,
        "couleur": jour_color(jour),
        "popup": {
            "action": field("Action"),
            "type": field("Type"),
            "billet": field("Billet"),
            "prix": field("Prix"),
            "city_card": field("City Card"),
            "ouverture": field("Ouverture"),
            "fermeture": field("Fermeture"),
            "remarque": field("Remarque"),
        },
    }


def find_ordre_collisions(wb: openpyxl.Workbook) -> dict[tuple[int, int], list[str]]:
    seen: dict[tuple[int, int], list[str]] = {}
    for item in iter_activity_rows(wb):
        seen.setdefault((item["jour"], item["visite"]), []).append(item["nom"])
    return {key: noms for key, noms in seen.items() if len(noms) > 1}


def find_duplicate_ordre_labels(wb: openpyxl.Workbook) -> dict[str, list[str]]:
    seen: dict[str, list[str]] = {}
    for item in iter_activity_rows(wb):
        label = item["ordre_label"]
        seen.setdefault(label, []).append(f"{item['sheet_name']}:{item['nom']}")
    return {label: locs for label, locs in seen.items() if len(locs) > 1}


def build_voyage_data(wb: openpyxl.Workbook) -> dict[str, Any]:
    points: list[dict[str, Any]] = []
    jours: set[int] = set()

    for item in iter_activity_rows(wb):
        jours.add(item["jour"])
        point = row_to_point(item)
        if point:
            points.append(point)

    points.sort(key=lambda p: (p["jour"], p["visite"], p["id"]))
    return {"jours": sorted(jours), "points": points}
