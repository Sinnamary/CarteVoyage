#!/usr/bin/env python3
"""Géocode les lieux du classeur de planning et remplit Latitude/Longitude."""

from __future__ import annotations

import argparse
import csv
import json
import sys
import time
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
import requests

from excel_utils import (
    backup_excel,
    cell_value,
    data_dir,
    day_sheets,
    default_excel_path,
    ensure_map_columns,
    has_coordinates,
    iter_activity_rows,
    normalize_text,
    ville_for_row,
)

NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
USER_AGENT = "CarteVoyage/1.0 (voyage planning map)"
REQUEST_DELAY = 1.1

MANUAL_COORDS = {
    "allard pierson": (52.368944, 4.892944),
    "Allard Pierson Museum Amsterdam": (52.368944, 4.892944),
    "joods historisch museum": (52.367015, 4.903445),
    "Joods Historisch Museum Amsterdam": (52.367015, 4.903445),
    "Römisch-Germanisches Museum Köln": (50.9422, 6.9633),
    "Kolumba Kunstmuseum Köln": (50.9364, 6.9556),
    "Brauhaus Päffgen Köln": (50.9488, 6.9423),
    "Ennevelin France": (50.5286, 3.2278),
    "Cap Gris-Nez": (50.8625, 1.7444),
    "Côte d'Opale": (50.7259, 1.6137),
    "côte d'opale": (50.7259, 1.6137),
    "Côte d'Opale Boulogne-sur-Mer France": (50.7259, 1.6137),
    "Vieille Bourse de Lille": (50.6370472, 3.0640170),
    "Vieille Bourse Lille": (50.6370472, 3.0640170),
    "Ville de Bergues": (50.9683886, 2.4325247),
    "Bergues": (50.9683886, 2.4325247),
    "Bergues, Nord, France": (50.9683886, 2.4325247),
}

GENERIC_REMARQUES = frozenset(
    {"centre", "center", "centrum", "nord", "sud", "est", "ouest", "—", "-", ""}
)

# Noms officiels dans la colonne Excel « Lieu » (anciens noms ambigus -> noms exacts).
EXCEL_LIEU_RENAMES: dict[str, str] = {
    "Allard Pierson": "Allard Pierson Museum, Amsterdam",
    "Anne Frank Huis": "Anne Frank Huis, Amsterdam",
    "Begijnhof": "Begijnhof, Amsterdam",
    "Bob W Amsterdam Noord": "Bob W Amsterdam Noord, Amsterdam",
    "Capr Gris nez": "Cap Gris-Nez",
    "Cathédrale": "Cathédrale de Cologne (Kölner Dom)",
    "Cote opale": "Côte d'Opale",
    "Het Grachtenhuis": "Het Grachtenhuis, Amsterdam",
    "Het Scheepvaart Museum": "Het Scheepvaartmuseum, Amsterdam",
    "Houseboat Museum": "Houseboat Museum, Amsterdam",
    "Hortus Botanicus": "Hortus Botanicus, Amsterdam",
    "Joods Historisch Museum": "Joods Historisch Museum, Amsterdam",
    "Kolumba": "Kolumba Kunstmuseum, Köln",
    "Micropia": "ARTIS Micropia, Amsterdam",
    "Museum Ludwig": "Museum Ludwig, Köln",
    "National Holocaust Museum": "National Holocaust Museum, Amsterdam",
    "Rembrandthuis": "Museum Rembrandthuis, Amsterdam",
    "Rijksmuseum": "Rijksmuseum, Amsterdam",
    "Römisch-Germanisches Museum": "Römisch-Germanisches Museum, Köln",
    "Rudi's Stroopwaffels": "Rudi's Original Stroopwafels, Amsterdam",
    "Stedelijk Museum": "Stedelijk Museum, Amsterdam",
    "Van Gogh Museum": "Van Gogh Museum, Amsterdam",
    "Verzetmuseum": "Verzetsmuseum, Amsterdam",
    "Vieille Bourse (Lille)": "Vieille Bourse de Lille",
    "Wereldmuseum": "Wereldmuseum, Amsterdam",
    "Amsterdam Museum": "Amsterdam Museum, Amsterdam",
    "De Bakkerswinkel": "De Bakkerswinkel, Amsterdam",
    "Pompadour": "Pompadour, Amsterdam",
}

NAME_ALIASES: dict[str, str] = {
    # Typo / variantes Excel
    "Het Bejinhof": "Begijnhof Amsterdam",
    "Risjkmuseum": "Rijksmuseum Amsterdam",
    "Rembrandhuis": "Museum Rembrandthuis Amsterdam",
    "Croisière sur les canaux": "Amsterdam canal cruise Centrum",
    "Balade dans l'ancien quartier juif": "Jodenbuurt Amsterdam",
    "Les grands canaux": "Grachtengordel Amsterdam",
    "Tour A'DAM": "A'DAM Lookout Amsterdam",
    "Départ Strasbourg": "Strasbourg Robertsau",
    "Cap Gris-Nez": "Cap Gris-Nez France",
    "Côte d'Opale": "Côte d'Opale Boulogne-sur-Mer France",
    "Brauhaus Päffgen Friesenstraße 64–66, 50670 Köln": "Brauhaus Päffgen Köln",
    # Noms Excel officiels -> requete Nominatim sans ambiguite
    "Allard Pierson Museum, Amsterdam": "Allard Pierson Museum Amsterdam",
    "Anne Frank Huis, Amsterdam": "Anne Frank House Amsterdam",
    "ARTIS Micropia, Amsterdam": "ARTIS Micropia Amsterdam",
    "Begijnhof, Amsterdam": "Begijnhof Amsterdam",
    "Bob W Amsterdam Noord, Amsterdam": "Bob W Amsterdam Noord Amsterdam",
    "Cathédrale de Cologne (Kölner Dom)": "Kölner Dom",
    "Het Grachtenhuis, Amsterdam": "Het Grachtenhuis Amsterdam",
    "Het Scheepvaartmuseum, Amsterdam": "Het Scheepvaartmuseum Amsterdam",
    "Houseboat Museum, Amsterdam": "Houseboat Museum Amsterdam",
    "Hortus Botanicus, Amsterdam": "Hortus Botanicus Amsterdam",
    "Joods Historisch Museum, Amsterdam": "Joods Historisch Museum Amsterdam",
    "Kolumba Kunstmuseum, Köln": "Kolumba Kunstmuseum Köln",
    "Museum Ludwig, Köln": "Museum Ludwig Köln",
    "Museum Rembrandthuis, Amsterdam": "Museum Rembrandthuis Amsterdam",
    "National Holocaust Museum, Amsterdam": "National Holocaust Museum Amsterdam",
    "Rijksmuseum, Amsterdam": "Rijksmuseum Amsterdam",
    "Römisch-Germanisches Museum, Köln": "Römisch-Germanisches Museum Köln",
    "Rudi's Original Stroopwafels, Amsterdam": "Rudi's Original Stroopwafels Amsterdam",
    "Stedelijk Museum, Amsterdam": "Stedelijk Museum Amsterdam",
    "Van Gogh Museum, Amsterdam": "Van Gogh Museum Amsterdam",
    "Verzetsmuseum, Amsterdam": "Verzetsmuseum Amsterdam",
    "Vieille Bourse de Lille": "Vieille Bourse Lille",
    "Ville de Bergues": "Bergues, Nord, France",
    "Wereldmuseum, Amsterdam": "Wereldmuseum Amsterdam",
    "Amsterdam Museum, Amsterdam": "Amsterdam Museum Amsterdam",
    "De Bakkerswinkel, Amsterdam": "De Bakkerswinkel Amsterdam Centrum",
    "Pompadour, Amsterdam": "Pompadour Amsterdam",
    # Noms courts encore presents ou sans ville
    "Allard Pierson": "Allard Pierson Museum Amsterdam",
    "Amsterdam Museum": "Amsterdam Museum Amsterdam",
    "Begijnhof": "Begijnhof Amsterdam",
    "De Bakkerswinkel": "De Bakkerswinkel Amsterdam Centrum",
    "Het Grachtenhuis": "Het Grachtenhuis Amsterdam",
    "Hotel Pullman Cologne": "Pullman Cologne Hotel",
    "Houseboat Museum": "Houseboat Museum Amsterdam",
    "Joods Historisch Museum": "Joods Historisch Museum Amsterdam",
    "Jordaan": "Jordaan Amsterdam",
    "Kolumba": "Kolumba Kunstmuseum Köln",
    "Micropia": "ARTIS Micropia Amsterdam",
    "Museum Ludwig": "Museum Ludwig Köln",
    "Oostelijk havengebied": "Oostelijk Havengebied Amsterdam",
    "Pompadour": "Pompadour Amsterdam",
    "Römisch-Germanisches Museum": "Römisch-Germanisches Museum Köln",
    "Vondelpark": "Vondelpark Amsterdam",
}

COUNTRY_BY_VILLE: dict[str, str] = {
    "amsterdam": "nl",
    "cologne": "de",
    "köln": "de",
    "lille": "fr",
    "strasbourg": "fr",
    "ennevelin": "fr",
    "cap gris-nez": "fr",
    "bergues": "fr",
    "boulogne-sur-mer": "fr",
    "côte d'opale": "fr",
    "cote d'opale": "fr",
}

NOM_ALIASES_BY_VILLE: dict[tuple[str, str], str] = {
    ("cathédrale", "cologne"): "Kölner Dom",
    ("cathédrale de cologne (kölner dom)", "cologne"): "Kölner Dom",
    ("cathédrale", "köln"): "Kölner Dom",
}

# Villes manquantes dans Excel a completer selon le lieu.
EXCEL_VILLE_BY_NOM: dict[str, str] = {
    "Bob W Amsterdam Noord": "Amsterdam",
    "Bob W Amsterdam Noord, Amsterdam": "Amsterdam",
    "Kolumba": "Cologne",
    "Kolumba Kunstmuseum, Köln": "Cologne",
    "Römisch-Germanisches Museum": "Cologne",
    "Römisch-Germanisches Museum, Köln": "Cologne",
    "Ennevelin": "Ennevelin",
    "Cap Gris-Nez": "Cap Gris-Nez",
    "Côte d'Opale": "Boulogne-sur-Mer",
    "Ville de Bergues": "Bergues",
}


def load_cache(cache_path: Path) -> dict:
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    return {}


def save_cache(cache_path: Path, cache: dict) -> None:
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def resolve_nom(nom: str) -> str:
    return NAME_ALIASES.get(nom, nom)


def resolve_nom_for_ville(nom: str, ville: str) -> str:
    key = (normalize_text(nom).lower(), normalize_text(ville).lower())
    if key in NOM_ALIASES_BY_VILLE:
        return NOM_ALIASES_BY_VILLE[key]
    if nom in NAME_ALIASES:
        return NAME_ALIASES[nom]
    if nom.lower() in NAME_ALIASES:
        return NAME_ALIASES[nom.lower()]
    return resolve_nom(nom)


def country_for_ville(ville: str) -> str:
    return COUNTRY_BY_VILLE.get(ville.strip().lower(), "nl")


def is_non_geocodable_lieu(nom: str) -> bool:
    """Trajets logistiques sans point sur la carte."""
    lower = normalize_text(nom).lower()
    return lower.startswith(("trajet ", "retour "))


def build_queries(nom: str, remarque: str, action: str = "", ville: str = "") -> list[str]:
    search_name = resolve_nom(nom)
    queries: list[str] = []
    remarque_key = normalize_text(remarque).lower()

    if remarque and action.lower() == "balade":
        if ville:
            queries.append(f"{remarque}, {ville}")
        elif remarque_key not in GENERIC_REMARQUES:
            queries.append(remarque)
    queries.append(search_name)
    if remarque and remarque_key not in GENERIC_REMARQUES:
        queries.append(f"{search_name}, {remarque}")
    if ville:
        queries.append(f"{search_name}, {ville}")

    seen: set[str] = set()
    unique: list[str] = []
    for q in queries:
        q = q.strip(" ,")
        if q and q not in seen:
            seen.add(q)
            unique.append(q)
    return unique


def nominatim_search(query: str, country: str) -> tuple[float, float] | None:
    params = {
        "q": query,
        "format": "json",
        "limit": 1,
        "countrycodes": country,
    }
    response = requests.get(
        NOMINATIM_URL,
        params=params,
        headers={"User-Agent": USER_AGENT},
        timeout=30,
    )
    response.raise_for_status()
    results = response.json()
    if not results:
        return None
    return float(results[0]["lat"]), float(results[0]["lon"])


def geocode_place(
    nom: str,
    remarque: str,
    cache: dict,
    country: str,
    action: str = "",
    use_cache: bool = True,
    ville: str = "",
) -> tuple[float, float] | None:
    search_nom = resolve_nom_for_ville(nom, ville)
    if search_nom in MANUAL_COORDS:
        return MANUAL_COORDS[search_nom]
    if nom in MANUAL_COORDS:
        return MANUAL_COORDS[nom]
    if nom.lower() in MANUAL_COORDS:
        return MANUAL_COORDS[nom.lower()]
    if search_nom.lower() in MANUAL_COORDS:
        return MANUAL_COORDS[search_nom.lower()]

    cache_key = f"{country}|{search_nom}|{remarque}"
    if use_cache and cache_key in cache:
        entry = cache[cache_key]
        return entry["lat"], entry["lon"]

    for query in build_queries(search_nom, remarque, action, ville):
        query_key = f"query|{country}|{query}"
        if use_cache and query_key in cache:
            entry = cache[query_key]
            coords = (entry["lat"], entry["lon"])
            cache[cache_key] = {"lat": coords[0], "lon": coords[1], "source": "cache"}
            return coords

        time.sleep(REQUEST_DELAY)
        try:
            coords = nominatim_search(query, country)
        except requests.RequestException as exc:
            print(f"WARN requete Nominatim en echec pour '{query}': {exc}")
            coords = None

        if coords:
            cache[query_key] = {"lat": coords[0], "lon": coords[1], "source": "nominatim"}
            cache[cache_key] = {
                "lat": coords[0],
                "lon": coords[1],
                "source": "nominatim",
                "query": query,
            }
            return coords

    return None


def normalize_lieu(nom: str) -> str:
    return normalize_text(nom).replace("\xa0", " ")


def apply_excel_lieu_renames(wb: openpyxl.Workbook) -> int:
    """Renomme les lieux ambigus et complete les villes manquantes dans Excel."""
    updated = 0
    for item in iter_activity_rows(wb):
        nom = normalize_lieu(item["nom"])
        col_index = item["col_index"]
        ws = item["ws"]
        row_idx = item["row_idx"]
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))

        new_nom = EXCEL_LIEU_RENAMES.get(nom)
        if new_nom and new_nom != nom:
            ws.cell(row=row_idx, column=col_index["Nom"] + 1, value=new_nom)
            updated += 1
            nom = new_nom

        if "Ville" in col_index:
            ville = normalize_text(cell_value(row, col_index, "Ville"))
            if not ville and nom in EXCEL_VILLE_BY_NOM:
                ws.cell(row=row_idx, column=col_index["Ville"] + 1, value=EXCEL_VILLE_BY_NOM[nom])
                updated += 1

    return updated


def ensure_columns(wb: openpyxl.Workbook) -> None:
    for sheet_name in day_sheets(wb):
        ensure_map_columns(wb[sheet_name])


def run_geocoding(excel_path: Path, dry_run: bool = False, force: bool = False) -> None:
    wb = openpyxl.load_workbook(excel_path)
    ensure_columns(wb)

    renamed = 0
    if not dry_run:
        renamed = apply_excel_lieu_renames(wb)
        if renamed:
            print(f"Excel: {renamed} champ(s) mis a jour (noms / villes)")

    cache_path = data_dir() / "geocode_cache.json"
    cache = load_cache(cache_path)
    errors: list[dict[str, str]] = []
    updated = 0
    skipped = 0

    for item in iter_activity_rows(wb):
        col_index = item["col_index"]
        ws = item["ws"]
        row_idx = item["row_idx"]
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))

        nom = normalize_text(cell_value(row, col_index, "Nom"))

        if is_non_geocodable_lieu(nom):
            skipped += 1
            continue

        if has_coordinates(row, col_index) and not force:
            skipped += 1
            continue
        remarque = normalize_text(cell_value(row, col_index, "Remarque"))
        action = normalize_text(cell_value(row, col_index, "Action"))
        ville = ville_for_row(row, col_index)
        country = country_for_ville(ville)

        coords = geocode_place(
            nom,
            remarque,
            cache,
            country=country,
            action=action,
            use_cache=not force,
            ville=ville,
        )

        label = item["ordre_label"]
        if coords:
            if not dry_run:
                ws.cell(row=row_idx, column=col_index["Latitude"] + 1, value=coords[0])
                ws.cell(row=row_idx, column=col_index["Longitude"] + 1, value=coords[1])
            updated += 1
            print(f"OK  [{label}] {nom} ({ville or country}) -> {coords[0]:.6f}, {coords[1]:.6f}")
        else:
            errors.append(
                {
                    "jour": str(item["jour"]),
                    "visite": str(item["visite"]),
                    "nom": nom,
                    "ville": ville,
                    "pays": country,
                    "requete": build_queries(nom, remarque, action, ville)[0],
                }
            )
            print(f"ERR [{label}] {nom} -> non trouve")

    save_cache(cache_path, cache)

    errors_path = data_dir() / "geocode_errors.csv"
    with errors_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=["jour", "visite", "nom", "ville", "pays", "requete"]
        )
        writer.writeheader()
        writer.writerows(errors)

    if not dry_run and (updated > 0 or renamed):
        backup_path = backup_excel(excel_path)
        wb.save(excel_path)
        print(f"Backup: {backup_path}")

    print(f"\nTermine: {updated} geocode(s), {skipped} ignore(s), {len(errors)} erreur(s)")
    print(f"Cache: {cache_path}")
    print(f"Erreurs: {errors_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Geocode les lieux dans le fichier Excel de voyage.")
    parser.add_argument("excel", nargs="?", default=str(default_excel_path()))
    parser.add_argument("--dry-run", action="store_true", help="Simuler sans ecrire dans Excel")
    parser.add_argument("--force", action="store_true", help="Re-geocoder meme si Lat/Lon presentes")
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    if not excel_path.exists():
        raise SystemExit(f"Fichier introuvable: {excel_path}")

    run_geocoding(excel_path, dry_run=args.dry_run, force=args.force)


if __name__ == "__main__":
    main()
