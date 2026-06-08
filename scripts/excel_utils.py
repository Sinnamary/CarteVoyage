"""Utilitaires partagés pour lire/écrire le fichier Excel de voyage."""

from __future__ import annotations

import re
import shutil
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

PLANNING_SHEETS = {"Amsterdam"}
CITY_BY_SHEET = {
    "Cologne": "Cologne",
    "Lille": "Lille",
}
AMSTERDAM_QUARTIER_SHEETS = {
    "Centre - Jordaan",
    "Vondelpark",
    "Est - Sud Est",
    "Nord",
}

BASE_COLUMNS = [
    "Action",
    "Nom",
    "Type",
    "Billet",
    "Prix",
    "City Card",
    "Ouverture",
    "Fermeture",
    "Remarque",
    "Site",
]
MAP_COLUMNS = ["Ordre", "Latitude", "Longitude", "Lien"]

SHEET_COLORS = {
    "Centre - Jordaan": "#e74c3c",
    "Vondelpark": "#27ae60",
    "Est - Sud Est": "#3498db",
    "Nord": "#9b59b6",
    "Cologne": "#e67e22",
    "Lille": "#1abc9c",
}


def project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def excel_dir() -> Path:
    path = project_root() / "excel"
    path.mkdir(exist_ok=True)
    return path


def data_dir() -> Path:
    path = project_root() / "data"
    path.mkdir(exist_ok=True)
    return path


def web_dir() -> Path:
    path = project_root() / "web"
    path.mkdir(exist_ok=True)
    return path


def default_excel_path() -> Path:
    return excel_dir() / "Voyage Amsterdam.xlsx"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sheet_city(sheet_name: str) -> str:
    if sheet_name in CITY_BY_SHEET:
        return CITY_BY_SHEET[sheet_name]
    if sheet_name in AMSTERDAM_QUARTIER_SHEETS:
        return "Amsterdam"
    return sheet_name


def is_activity_sheet(sheet_name: str) -> bool:
    return sheet_name not in PLANNING_SHEETS


def normalize_sheet_key(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def match_sheet_name(label: str, sheet_names: list[str]) -> str | None:
    label_key = normalize_sheet_key(label)
    for name in sheet_names:
        if normalize_sheet_key(name) == label_key:
            return name
    for name in sheet_names:
        name_key = normalize_sheet_key(name)
        if label_key in name_key or name_key in label_key:
            return name
    return None


def parse_planning_days(wb: openpyxl.Workbook) -> dict[int, list[str]]:
    if "Amsterdam" not in wb.sheetnames:
        return {}

    ws = wb["Amsterdam"]
    day_map: dict[int, list[str]] = {}
    day_num = 0

    for row in ws.iter_rows(values_only=True):
        cells = [normalize_text(c) for c in row]
        if not any(cells):
            continue

        joined = " ".join(c for c in cells if c).lower()
        if "journ" in joined or re.search(r"\bjour\b", joined):
            day_num += 1

        for cell in cells:
            if not cell:
                continue
            lower = cell.lower()
            if "journ" in lower or re.search(r"\bjour\b", lower):
                continue
            matched = match_sheet_name(cell, wb.sheetnames)
            if matched and is_activity_sheet(matched):
                day_map.setdefault(day_num, [])
                if matched not in day_map[day_num]:
                    day_map[day_num].append(matched)

    return day_map


def find_header_row(ws: Worksheet) -> int | None:
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [normalize_text(c) for c in row]
        if "Nom" in values:
            return row_idx
    return None


def build_column_index(header_row: tuple[Any, ...]) -> dict[str, int]:
    index: dict[str, int] = {}
    for col_idx, value in enumerate(header_row):
        name = normalize_text(value)
        if name:
            index[name] = col_idx
    return index


def ensure_map_columns(ws: Worksheet, header_row_idx: int) -> dict[str, int]:
    header = [cell.value for cell in ws[header_row_idx]]
    col_index = build_column_index(tuple(header))

    next_col = len(header) + 1
    for col_name in MAP_COLUMNS:
        if col_name not in col_index:
            ws.cell(row=header_row_idx, column=next_col, value=col_name)
            col_index[col_name] = next_col - 1
            next_col += 1
        else:
            next_col = max(next_col, col_index[col_name] + 2)

    return col_index


def cell_value(row: tuple[Any, ...], col_index: dict[str, int], column: str) -> Any:
    if column not in col_index:
        return None
    idx = col_index[column]
    if idx >= len(row):
        return None
    return row[idx]


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def has_coordinates(row: tuple[Any, ...], col_index: dict[str, int]) -> bool:
    lat = parse_float(cell_value(row, col_index, "Latitude"))
    lon = parse_float(cell_value(row, col_index, "Longitude"))
    return lat is not None and lon is not None


def backup_excel(excel_path: Path) -> Path:
    backup_dir = excel_path.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    backup_path = backup_dir / (excel_path.stem + ".backup" + excel_path.suffix)
    shutil.copy2(excel_path, backup_path)
    return backup_path


def iter_activity_rows(wb: openpyxl.Workbook):
    day_map = parse_planning_days(wb)
    sheet_to_day: dict[str, int] = {}
    for day, sheets in day_map.items():
        for sheet in sheets:
            sheet_to_day[sheet] = day

    for sheet_name in wb.sheetnames:
        if not is_activity_sheet(sheet_name):
            continue

        ws = wb[sheet_name]
        header_row_idx = find_header_row(ws)
        if header_row_idx is None:
            continue

        header = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))
        col_index = build_column_index(header)
        if "Nom" not in col_index:
            continue

        ville = sheet_city(sheet_name)
        jour = sheet_to_day.get(sheet_name)
        row_counter = 0

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            nom = normalize_text(cell_value(row, col_index, "Nom"))
            if not nom:
                continue

            row_counter += 1
            ordre = parse_int(cell_value(row, col_index, "Ordre")) or row_counter

            yield {
                "sheet_name": sheet_name,
                "row_idx": row_idx,
                "ville": ville,
                "jour": jour,
                "ordre": ordre,
                "nom": nom,
                "col_index": col_index,
                "row": row,
                "ws": ws,
                "header_row_idx": header_row_idx,
            }


def row_to_point(item: dict[str, Any]) -> dict[str, Any] | None:
    row = item["row"]
    col_index = item["col_index"]

    lat = parse_float(cell_value(row, col_index, "Latitude"))
    lon = parse_float(cell_value(row, col_index, "Longitude"))
    if lat is None or lon is None:
        return None

    lien = normalize_text(cell_value(row, col_index, "Lien"))
    site = normalize_text(cell_value(row, col_index, "Site"))
    url = lien or site or None

    def field(name: str) -> Any:
        value = cell_value(row, col_index, name)
        if value is None or value == "":
            return None
        return value

    sheet_name = item["sheet_name"]
    return {
        "id": f"{sheet_name}-{item['row_idx']}",
        "ordre": item["ordre"],
        "jour": item["jour"],
        "ville": item["ville"],
        "onglet": sheet_name,
        "nom": item["nom"],
        "lat": lat,
        "lon": lon,
        "lien": url,
        "couleur": SHEET_COLORS.get(sheet_name, "#3388ff"),
        "popup": {
            "action": field("Action"),
            "type": field("Type"),
            "billet": field("Billet"),
            "prix": field("Prix"),
            "city_card": field("City Card"),
            "ouverture": field("Ouverture"),
            "fermeture": field("Fermeture"),
            "remarque": field("Remarque"),
        },
    }


def build_voyage_data(wb: openpyxl.Workbook) -> dict[str, Any]:
    day_map = parse_planning_days(wb)
    points: list[dict[str, Any]] = []
    onglets: list[str] = []
    villes: set[str] = set()

    for item in iter_activity_rows(wb):
        onglet = item["sheet_name"]
        if onglet not in onglets:
            onglets.append(onglet)
        villes.add(item["ville"])

        point = row_to_point(item)
        if point:
            points.append(point)

    return {
        "villes": sorted(villes),
        "onglets": onglets,
        "jours": {str(k): v for k, v in day_map.items()},
        "points": points,
    }
