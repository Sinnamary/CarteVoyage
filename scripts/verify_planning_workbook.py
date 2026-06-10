"""Verification du classeur de planning."""

from __future__ import annotations

import openpyxl

from excel_utils import (
    DEFAULT_EXCEL_NAME,
    PLANNING_COLUMNS,
    build_listes_ranges,
    day_sheets,
    default_excel_path,
    excel_dir,
    find_duplicate_ordre_labels,
    find_ordre_collisions,
    iter_activity_rows,
    normalize_text,
    row_to_point,
    sync_listes_validations,
)

REQUIRED_OVERVIEW_MARKERS = ("Strasbourg", "Cologne", "Amsterdam", "Lille", "11 au 14 août 2026")
MAP_COLUMNS = ("Latitude", "Longitude")


def check_ordre_text_format(wb: openpyxl.Workbook) -> list[str]:
    issues: list[str] = []
    wb_fmt = openpyxl.load_workbook(default_excel_path(), data_only=False)
    for sheet_name in day_sheets(wb):
        ws = wb_fmt[sheet_name]
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(r, 1)
            etape = cell.value
            if etape is None:
                continue
            text = str(etape)
            if ".10" in text and cell.number_format != "@":
                issues.append(
                    f"{sheet_name} L{r}: {text!r} doit être en texte (@), format={cell.number_format!r}"
                )
    return issues


def main() -> None:
    path = excel_dir() / DEFAULT_EXCEL_NAME
    if not path.exists():
        raise SystemExit(f"Fichier introuvable: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    errors: list[str] = []
    warnings: list[str] = []

    expected_meta = ["Vue d'ensemble", "Listes"] + [f"Jour {d}" for d in range(1, 13)]
    if wb.sheetnames != expected_meta:
        errors.append(f"Feuilles inattendues: {wb.sheetnames}")

    listes = wb["Listes"]
    if listes.sheet_state != "hidden":
        warnings.append("Feuille Listes devrait être masquée")

    wb_validations = openpyxl.load_workbook(path)
    expected_ranges = build_listes_ranges(wb_validations)
    pending = sync_listes_validations(wb_validations)
    if pending:
        warnings.append(
            "Listes deroulantes desynchronisees: lancer python scripts/sync_listes_validations.py "
            f"(ex. Ville attendue {expected_ranges.get('E', '?')})"
        )

    overview = wb["Vue d'ensemble"]
    overview_text = " ".join(
        str(cell.value) for row in overview.iter_rows() for cell in row if cell.value
    )
    for marker in REQUIRED_OVERVIEW_MARKERS:
        if marker not in overview_text:
            warnings.append(f"Vue d'ensemble: '{marker}' absent")

    for sheet_name in day_sheets(wb):
        ws = wb[sheet_name]
        header_names = {
            normalize_text(ws.cell(2, c).value)
            for c in range(1, ws.max_column + 1)
            if ws.cell(2, c).value
        }
        missing = [col for col in PLANNING_COLUMNS.values() if col not in header_names]
        if missing:
            errors.append(f"{sheet_name}: colonnes manquantes {missing}")

        missing_map = [col for col in MAP_COLUMNS if col not in header_names]
        if missing_map:
            warnings.append(
                f"{sheet_name}: {', '.join(missing_map)} absentes (ajoutées par geocode_excel.py)"
            )

        banner = ws.cell(1, 1).value or ""
        if not banner.startswith(sheet_name):
            warnings.append(f"{sheet_name}: bannière inattendue ({banner[:50]}...)")

    for (jour, visite), noms in sorted(find_ordre_collisions(wb).items()):
        warnings.append(f"Collision ordre {jour}.{visite}: {', '.join(noms)}")

    for label, locs in sorted(find_duplicate_ordre_labels(wb).items()):
        warnings.append(f"N° étape {label} répété: {', '.join(locs)}")

    errors.extend(check_ordre_text_format(wb))

    items = list(iter_activity_rows(wb))
    with_coords = sum(1 for item in items if row_to_point(item))

    print(f"Fichier: {path}")
    print(f"Lignes activité: {len(items)}")
    print(f"Avec coordonnées (carte): {with_coords}")
    print(f"Sans coordonnées: {len(items) - with_coords}")

    if warnings:
        print("\nAvertissements:")
        for w in warnings:
            print(f"  - {w}")

    if errors:
        print("\nProblèmes:")
        for e in errors:
            print(f"  - {e}")
        raise SystemExit(1)

    print("\nOK: classeur valide")


if __name__ == "__main__":
    main()
